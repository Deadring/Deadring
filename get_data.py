from datetime import datetime
import json
from pathlib import Path

from openpyxl import load_workbook

PATH = Path(__file__).parent

sheets = ['tbPelanggan', 'tbSalesman', 'tbKasir', 'tbBarang', 'tbDetailFaktur', 'tbTransaksiBarang', 'tbDetailRetur', 'tbBarangRetur']

def get_data(excel_file=None):
    wb = load_workbook(excel_file, data_only=True)
    data = {}

    for sheet in sheets:
        record = list(wb[sheet].iter_rows(values_only=True))
        head = record[0]
        data[sheet] = {}

        for i, cell in enumerate(record[1:]):
            data[sheet][i] = {header: x.strftime('%Y-%m-%d') if isinstance((x:=cell[j]), datetime) else x for j, header in enumerate(head)}

    return data

if __name__ == '__main__':
    dict_data = get_data(PATH / 'A1.xlsx')

    with open(PATH / 'data.json', 'w+') as json_file:
        json.dump(dict_data, json_file)